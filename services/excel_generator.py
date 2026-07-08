import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from .logger import logger

def generate_excel_report(image_data: list, output_filename: str = "Image_Checkpoints.xlsx") -> str:
    """
    Generates a professional Excel report for the extracted image data.
    """
    logger.info(f"Generating Excel report with {len(image_data)} rows.")
    wb = Workbook()
    ws = wb.active
    ws.title = "Image Checkpoints"

    headers = [
        "No", "Tag", "Src", "Alt", 
        "DisplayedWidth", "DisplayedHeight", 
        "OriginalWidth", "OriginalHeight", 
        "Display Difference (%)",
        "Extension", "File Name", "Loading Attribute",
        "FetchPriority", "Decoding", "CrossOrigin", 
        "ReferrerPolicy", "Lazy Loaded", "Responsive", 
        "Is Duplicate", "Missing Alt", "Broken", 
        "Zero Dimensions", "Upscaled", "Downscaled", "Missing W/H Attributes"
    ]

    # Append header
    ws.append(headers)

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(wrap_text=True, vertical="center")

    # Apply header styles
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    ws.freeze_panes = "A2"

    # Append data
    for idx, data in enumerate(image_data, 1):
        row = [
            idx,
            data.get('tag', ''),
            data.get('src', ''),
            data.get('alt', ''),
            data.get('displayedWidth', ''),
            data.get('displayedHeight', ''),
            data.get('originalWidth', ''),
            data.get('originalHeight', ''),
            data.get('displayDifference', ''),
            data.get('extension', ''),
            data.get('fileName', ''),
            data.get('loadingAttribute', ''),
            data.get('fetchpriority', ''),
            data.get('decoding', ''),
            data.get('crossorigin', ''),
            data.get('referrerpolicy', ''),
            "Yes" if data.get('lazyLoaded') else "No",
            "Yes" if data.get('responsive') else "No",
            "Yes" if data.get('duplicate') else "No",
            "Yes" if data.get('missingAlt') else "No",
            "Yes" if data.get('broken') else "No",
            "Yes" if data.get('zeroDimensions') else "No",
            "Yes" if data.get('upscaled') else "No",
            "Yes" if data.get('downscaled') else "No",
            "Yes" if data.get('missingWHAttributes') else "No"
        ]
        ws.append(row)
        
        # Format the newly added row
        current_row = ws[ws.max_row]
        for cell_idx, cell in enumerate(current_row):
            cell.border = thin_border
            # Alternate row colors
            if idx % 2 == 0:
                cell.fill = alt_fill
            
            # Center alignment by default
            cell.alignment = center_align
            
            # Wrap text for Alt (column 4)
            if cell_idx == 3:
                cell.alignment = wrap_align
                
            # Hyperlink for Src (column 3)
            if cell_idx == 2 and data.get('src') and str(data.get('src')).startswith('http'):
                cell.hyperlink = data.get('src')
                cell.font = Font(color="0000FF", underline="single")

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Auto-size columns based on header or typical content
    col_widths = {
        "A": 5, "B": 10, "C": 50, "D": 40, 
        "E": 15, "F": 15, "G": 15, "H": 15, 
        "I": 20, "J": 10, "K": 20, "L": 15,
        "M": 15, "N": 15, "O": 15, "P": 15,
        "Q": 12, "R": 12, "S": 12, "T": 12,
        "U": 10, "V": 15, "W": 10, "X": 10, "Y": 20
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    downloads_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'downloads')
    os.makedirs(downloads_dir, exist_ok=True)
    
    file_path = os.path.join(downloads_dir, output_filename)
    wb.save(file_path)
    logger.info(f"Excel report saved to {file_path}")
    
    return file_path
